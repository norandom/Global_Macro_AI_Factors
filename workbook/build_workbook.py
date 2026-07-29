"""Headless openpyxl generator for the factor workbook skeleton (R7.4).

Emits ``factor_workbook.xlsx``: a navigation ``Index`` sheet plus the six
storyboard step sheets S0-S5. Each step sheet carries its title, the mandated
framing block verbatim (R2.2, R3.3, R6.3, R7.3), and formula cells that
reference only the worksheet functions exposed by ``factor_workbook.addin``
(``FW_LOAD``, ``FW_STEP``, ``FW_TABLE``, ``FW_CHECKS``, ``FW_FRAMING``,
``FW_PROVENANCE``, ``FW_VERSION``). The release tag lives once, in the named
cell ``RELEASE_TAG`` (``Index!$B$1``); the client handle produced by
``FW_LOAD`` lives once in ``Index!$B$2`` and every step sheet derives its
step-view handle from it. Generation is deterministic and needs no display,
so it runs on the Linux development machine (R7.4); the formulas only
evaluate under PyXLL in Excel.

Tag selection (task 10.7, R7.3-R8.8): the generated thesis workbook stays
pinned to ``data-v2`` — its S0-S5 walkthrough is an immutable historical
audit. ``data-v4`` is an explicit user selection (``generate(path,
tag="data-v4")`` or ``--tag data-v4``); the environment is never consulted,
so no variable can silently flip the default. Every selection constructs a
FRESH release client (:func:`release_client`), the on-disk cache stays keyed
by ``(tag, asset)`` inside the client, and ``data-v4`` loads run through the
manifest-verified path whose recorded Provenance exposes the active tag,
publication identity, manifest status, source URLs, and one row per asset.
"""

from __future__ import annotations

import argparse
import sys
from collections.abc import Callable
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.worksheet import Worksheet

from factor_workbook import steps
from factor_workbook.release import ReleaseClient

#: Default release tag pre-filled in ``Index!$B$1``: data-v2 is a superset of
#: v1 and carries the S0 static assets (task 7.1).
_DEFAULT_TAG = "data-v2"

#: The corrected release tag; opt-in only, never a silent default (task 10.7).
_DATA_V4_TAG = "data-v4"

_DATA_V4_NOTE = (
    "data-v4 is an explicit selection — the generated default stays data-v2. "
    "Every data-v4 load is manifest-verified before caching, and the "
    "provenance block below exposes the active tag, publication identity "
    "(verification), manifest status (verified / expected_sha256), source "
    "URLs, and one provenance row per loaded asset."
)

#: Sheet titles, mirroring the ``StepView.title`` strings the builders emit.
_TITLES = {
    "S0": "S0 — Static buy-and-hold line (hindsight-selected, in-sample)",
    "S1": "S1 — Model certification (no-recall screen)",
    "S2": "S2 — Coin-flip naive prediction",
    "S3": "S3 — AI macro-factor development (recall-guarded)",
    "S4": "S4 — Two-line walk-forward simulation",
    "S5": "S5 — Luck versus skill (contamination premium vs robust inference)",
}

#: Static named tables each step view exposes (per the committed builders).
#: S1 additionally serves DYNAMIC per-slug tables (``evidence:<slug>``,
#: ``class_stats:<slug>``); those get a documented drill-down note instead
#: of pre-placed formulas.
_TABLES = {
    "S0": ["equity_10y", "equity", "targets_drift", "stats", "crisis_episodes"],
    "S1": ["certification"],
    "S2": ["naive_eval", "summary"],
    "S3": [
        "loadings_v1",
        "scores_v1",
        "loadings_v2",
        "scores_v2",
        "score_summary",
        "stability",
        "views_v1",
        "gate",
    ],
    "S4": [
        "equity",
        "targets_pit",
        "detail_pit",
        "targets_nonpit",
        "detail_nonpit",
        "metrics",
    ],
    "S5": ["contrast", "premium", "ssr", "loading_stability"],
}

_S1_NOTE = (
    "Dynamic drill-down: per-slug S1 tables are served on demand — enter "
    '=FW_TABLE($B$1, "evidence:<slug>") or =FW_TABLE($B$1, "class_stats:<slug>") '
    "with a slug from the certification table."
)

# ponytail: fixed 40-row gap between table anchors; widen if a table outgrows it.
_TABLE_ROW_START = 8
_TABLE_ROW_STEP = 40

_WRAP = Alignment(wrap_text=True, vertical="top")
_BOLD = Font(bold=True)


def resolve_tag(tag: str | None = None) -> str:
    """Resolve the release tag by explicit selection only (task 10.7).

    ``None`` keeps the pinned ``data-v2`` default; any other version —
    including ``data-v4`` — must be passed explicitly. The environment is
    never consulted, so no variable can auto-upgrade the default. An explicit
    tag is validated through the release client's own immutable-tag rule.

    Raises:
        ValueError: For a mutable or unsafe explicit tag.
    """
    if tag is None:
        return _DEFAULT_TAG
    return ReleaseClient(tag).tag


def release_client(
    tag: str | None = None,
    cache_dir: Path | None = None,
    token_provider: Callable[[], str | None] | None = None,
) -> ReleaseClient:
    """Construct a FRESH release client for one explicitly selected tag.

    Every call returns a new client (R1.5): switching tags never reuses
    another client's state, and the client keys its on-disk cache by
    ``(tag, asset)``, so bytes cached under one tag are never served for
    another. Selecting ``data-v4`` routes every load through the
    manifest-verified task-10.2 path and records per-asset Provenance.
    """
    return ReleaseClient(
        resolve_tag(tag), cache_dir=cache_dir, token_provider=token_provider
    )


def _build_index(sheet: Worksheet, tag: str) -> None:
    """Navigation index: tag cell, client handle, provenance, step links."""
    sheet["A1"] = "Factor Workbook — Index"
    sheet["A1"].font = _BOLD
    sheet["B1"] = tag
    sheet["A2"] = "Client handle (edit B1 to re-load everything)"
    sheet["B2"] = "=FW_LOAD(Index!$B$1)"
    sheet["A3"] = "Release version"
    sheet["B3"] = "=FW_VERSION(Index!$B$2)"
    if tag == _DATA_V4_TAG:
        note = sheet["A4"]
        note.value = _DATA_V4_NOTE
        note.alignment = _WRAP
    sheet["A5"] = "Provenance"
    sheet["A5"].font = _BOLD
    sheet["B5"] = "=FW_PROVENANCE(Index!$B$2)"
    sheet["A14"] = "Storyboard steps"
    sheet["A14"].font = _BOLD
    for offset, step in enumerate(_TITLES):
        cell = sheet.cell(row=15 + offset, column=1, value=_TITLES[step])
        cell.hyperlink = Hyperlink(ref=cell.coordinate, location=f"'{step}'!A1")
        cell.font = Font(underline="single", color="0563C1")
    sheet.column_dimensions["A"].width = 60
    sheet.column_dimensions["B"].width = 40


def _build_step(sheet: Worksheet, step: str) -> None:
    """One step sheet: title, tag ref, framing, step/checks/table formulas."""
    sheet["A1"] = _TITLES[step]
    sheet["A1"].font = _BOLD
    sheet["A2"] = "='Index'!$B$1"
    framing = sheet["A3"]
    framing.value = getattr(steps, f"{step}_FRAMING")
    framing.alignment = _WRAP
    sheet.row_dimensions[3].height = 90
    sheet["B1"] = f'=FW_STEP(Index!$B$2, "{step}")'
    sheet["A4"] = "Framing (live)"
    sheet["B4"] = "=FW_FRAMING($B$1)"
    sheet["A5"] = "Checks"
    sheet["A5"].font = _BOLD
    sheet["B5"] = "=FW_CHECKS($B$1)"
    for offset, table in enumerate(_TABLES[step]):
        row = _TABLE_ROW_START + offset * _TABLE_ROW_STEP
        label = sheet.cell(row=row, column=1, value=table)
        label.font = _BOLD
        sheet.cell(row=row, column=2, value=f'=FW_TABLE($B$1, "{table}")')
    if step == "S1":
        note = sheet.cell(row=6, column=1, value=_S1_NOTE)
        note.alignment = _WRAP
    sheet.column_dimensions["A"].width = 60


def generate(path: str | Path, tag: str | None = None) -> Path:
    """Write the deterministic workbook skeleton to ``path`` and return it.

    Args:
        path: Destination ``.xlsx`` path.
        tag: Explicit release tag for ``Index!$B$1``; omitted keeps the
            pinned ``data-v2`` default (task 10.7).
    """
    workbook = Workbook()
    index = workbook.active
    index.title = "Index"
    _build_index(index, resolve_tag(tag))
    for step in _TITLES:
        _build_step(workbook.create_sheet(step), step)
    workbook.defined_names["RELEASE_TAG"] = DefinedName(
        "RELEASE_TAG", attr_text="Index!$B$1"
    )
    path = Path(path)
    workbook.save(path)
    return path


def main(argv: list[str] | None = None) -> int:
    """Generate the workbook; the tag comes only from an explicit ``--tag``."""
    parser = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    parser.add_argument(
        "--out",
        type=Path,
        default=Path(__file__).resolve().parent / "factor_workbook.xlsx",
        help="destination .xlsx path (default: factor_workbook.xlsx next to this script)",
    )
    parser.add_argument(
        "--tag",
        default=None,
        help=(
            "explicit release tag for Index!$B$1 (e.g. data-v4); omitted keeps "
            "the data-v2 default — the environment is never consulted"
        ),
    )
    args = parser.parse_args(argv)
    print(generate(args.out, tag=args.tag))
    return 0


if __name__ == "__main__":
    sys.exit(main())
