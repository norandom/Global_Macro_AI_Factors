"""Generated workbook skeleton: sheets, framing, and FW_* formula names (R7.4).

Opens the openpyxl-generated file headlessly and asserts the five step
sheets plus the navigation index exist, every ``=FW_...`` formula references
a function actually exposed by ``factor_workbook.addin`` (valid-name set
built by introspection, not a hardcoded list), and each sheet embeds its
mandated framing block verbatim (R2.2, R3.3, R6.3, R7.3).
"""

import re

import openpyxl
import pytest

import build_workbook
from factor_workbook import addin, steps

_STEP_SHEETS = ["S1", "S2", "S3", "S4", "S5"]
_FW_NAME = re.compile(r"FW_[A-Z_]+")


@pytest.fixture(scope="module")
def workbook(tmp_path_factory):
    """Generate the skeleton headlessly and load it with formulas intact."""
    path = tmp_path_factory.mktemp("wb") / "factor_workbook.xlsx"
    build_workbook.generate(path)
    return openpyxl.load_workbook(path, data_only=False)


def _formulas(workbook):
    """Every string cell value that is a formula, with its sheet name."""
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    yield sheet.title, cell.value


def test_sheets_are_index_plus_five_steps(workbook):
    assert workbook.sheetnames == ["Index"] + _STEP_SHEETS


def test_every_formula_references_an_exposed_function(workbook):
    exposed = {
        name.upper()
        for name, obj in vars(addin).items()
        if name.startswith("fw_") and callable(obj)
    }
    assert exposed, "introspection of factor_workbook.addin found no fw_* functions"
    seen = set()
    for _, formula in _formulas(workbook):
        for name in _FW_NAME.findall(formula):
            assert name in exposed, f"{formula!r} references unknown {name}"
            seen.add(name)
    assert seen == exposed, f"skeleton never uses {exposed - seen}"


def test_framing_blocks_embedded_verbatim(workbook):
    for step in _STEP_SHEETS:
        framing = getattr(steps, f"{step}_FRAMING")
        values = {c.value for row in workbook[step].iter_rows() for c in row}
        assert framing in values, f"{step} framing block missing"


def test_each_step_sheet_wires_step_checks_and_tables(workbook):
    for step in _STEP_SHEETS:
        per_sheet = [f for s, f in _formulas(workbook) if s == step]
        assert any(f'=FW_STEP(Index!$B$2, "{step}")' == f for f in per_sheet)
        assert any(f.startswith("=FW_CHECKS(") for f in per_sheet)
        assert any(f.startswith("=FW_TABLE(") for f in per_sheet)


def test_index_holds_tag_load_provenance_and_nav(workbook):
    index = workbook["Index"]
    assert index["B1"].value == "data-v1"
    assert "RELEASE_TAG" in workbook.defined_names
    formulas = [f for s, f in _formulas(workbook) if s == "Index"]
    assert "=FW_LOAD(Index!$B$1)" in formulas
    assert any(f.startswith("=FW_PROVENANCE(") for f in formulas)
    assert any(f.startswith("=FW_VERSION(") for f in formulas)
    links = {
        c.hyperlink.location
        for row in index.iter_rows()
        for c in row
        if c.hyperlink is not None
    }
    assert links == {f"'{s}'!A1" for s in _STEP_SHEETS}


def test_s1_documents_dynamic_drilldown_pattern(workbook):
    values = [
        c.value
        for row in workbook["S1"].iter_rows()
        for c in row
        if isinstance(c.value, str)
    ]
    assert any('FW_TABLE($B$1, "evidence:<slug>")' in v for v in values)
