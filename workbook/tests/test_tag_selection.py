"""Task 10.7: preserved ``data-v2`` default and explicit ``data-v4`` selection.

The generated thesis workbook stays pinned to ``data-v2``; ``data-v4`` is an
explicit user selection — never an environment-driven default flip. Explicit
selection constructs a FRESH release client per call, keeps on-disk caches
strictly separated by tag (switching tags never serves bytes cached under
another tag), and routes ``data-v4`` loads through the manifest-verified
task-10.2 path. All transport is mocked — no live network in the default run.
"""

import asyncio
import hashlib
import json

import openpyxl
import pytest

import build_workbook
from factor_workbook import addin, release
from factor_workbook.release import ReleaseError

DATA_V4_TAG = "data-v4"
ASSET = "portfolio_metrics_reader_ext2026.parquet"
MANIFEST = "publication_manifest.json"


def _url(tag: str, asset: str) -> str:
    return (
        "https://github.com/norandom/Global_Macro_AI_Factors/releases/download/"
        f"{tag}/{asset}"
    )


class FakeResponse:
    """Minimal stand-in for ``requests.Response``."""

    def __init__(self, status_code: int, content: bytes = b""):
        self.status_code = status_code
        self.content = content


def install_transport(monkeypatch, handler):
    """Replace ``requests.get`` with ``handler(url)``."""

    def fake_get(url, headers=None, timeout=None, **kwargs):
        return handler(url)

    monkeypatch.setattr(release.requests, "get", fake_get)


def make_data_v4_manifest(assets: dict[str, bytes]) -> bytes:
    """A completed publication manifest covering the given asset payloads."""
    return json.dumps(
        {
            "schema_id": "publication_manifest.v1",
            "release_tag": DATA_V4_TAG,
            "completed": True,
            "artifacts": [
                {
                    "path": name,
                    "sha256": hashlib.sha256(payload).hexdigest(),
                    "size": len(payload),
                }
                for name, payload in sorted(assets.items())
            ],
        },
        sort_keys=True,
    ).encode("utf-8")


def no_token() -> None:
    return None


def _tag_cell(path):
    return openpyxl.load_workbook(path, data_only=False)["Index"]["B1"].value


# --- data-v2 default preserved; no environment-based auto-upgrade ------------


def test_default_generation_stays_pinned_to_data_v2_despite_env(monkeypatch, tmp_path):
    for name in ("FACTOR_WORKBOOK_TAG", "FW_RELEASE_TAG", "RELEASE_TAG", "DATA_TAG"):
        monkeypatch.setenv(name, DATA_V4_TAG)

    assert build_workbook.resolve_tag(None) == "data-v2"
    assert _tag_cell(build_workbook.generate(tmp_path / "default.xlsx")) == "data-v2"


def test_main_cli_defaults_to_data_v2_and_requires_explicit_data_v4(monkeypatch, tmp_path, capsys):
    monkeypatch.setenv("FACTOR_WORKBOOK_TAG", DATA_V4_TAG)
    default_path = tmp_path / "default.xlsx"
    v4_path = tmp_path / "v4.xlsx"

    assert build_workbook.main(["--out", str(default_path)]) == 0
    assert build_workbook.main(["--out", str(v4_path), "--tag", DATA_V4_TAG]) == 0

    assert _tag_cell(default_path) == "data-v2"
    assert _tag_cell(v4_path) == DATA_V4_TAG
    printed = capsys.readouterr().out
    assert str(default_path) in printed and str(v4_path) in printed


# --- explicit data-v4 selection in the generated workbook --------------------


def test_generate_explicit_data_v4_keeps_provenance_surface_and_notes_selection(tmp_path):
    path = build_workbook.generate(tmp_path / "v4.xlsx", tag=DATA_V4_TAG)
    index = openpyxl.load_workbook(path, data_only=False)["Index"]

    assert index["B1"].value == DATA_V4_TAG
    formulas = {
        cell.value
        for row in index.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    }
    # active tag, per-asset provenance (url/sha/manifest status columns), and
    # the load handle all stay wired for the explicitly selected tag
    assert "=FW_LOAD(Index!$B$1)" in formulas
    assert "=FW_VERSION(Index!$B$2)" in formulas
    assert "=FW_PROVENANCE(Index!$B$2)" in formulas
    note = index["A4"].value
    assert isinstance(note, str) and "explicit" in note and "manifest" in note

    default_index = openpyxl.load_workbook(
        build_workbook.generate(tmp_path / "default.xlsx"), data_only=False
    )["Index"]
    assert default_index["B1"].value == "data-v2"
    assert default_index["A4"].value is None  # note only on explicit data-v4


def test_explicit_selection_rejects_mutable_tags(tmp_path):
    with pytest.raises(ValueError, match="immutable"):
        build_workbook.generate(tmp_path / "bad.xlsx", tag="latest")
    with pytest.raises(ValueError, match="immutable"):
        build_workbook.resolve_tag("current")


# --- fresh client per selection; no cross-tag cache substitution -------------


def test_release_client_factory_returns_a_fresh_client_per_selection(tmp_path):
    default = build_workbook.release_client(cache_dir=tmp_path, token_provider=no_token)
    v4 = build_workbook.release_client(DATA_V4_TAG, cache_dir=tmp_path, token_provider=no_token)
    v4_again = build_workbook.release_client(DATA_V4_TAG, cache_dir=tmp_path, token_provider=no_token)

    assert default.tag == "data-v2"
    assert v4.tag == DATA_V4_TAG
    assert v4 is not default and v4_again is not v4  # fresh client every call
    assert v4.provenance_table() == [] and v4_again.provenance_table() == []


def test_switching_tags_never_serves_bytes_cached_under_another_tag(monkeypatch, tmp_path):
    v2_payload = b"v2-cached-bytes"
    install_transport(monkeypatch, lambda url: FakeResponse(200, v2_payload))
    v2 = build_workbook.release_client("data-v2", cache_dir=tmp_path, token_provider=no_token)
    assert v2.fetch(ASSET)[0] == v2_payload
    assert (tmp_path / "data-v2" / ASSET).read_bytes() == v2_payload

    # switching to data-v4 with the release absent must fail, not substitute
    install_transport(monkeypatch, lambda url: FakeResponse(404))
    v4 = build_workbook.release_client(DATA_V4_TAG, cache_dir=tmp_path, token_provider=no_token)
    with pytest.raises(ReleaseError) as exc:
        v4.fetch(ASSET)
    assert exc.value.error.cause == "missing"
    assert v4.provenance_table() == []
    assert not (tmp_path / DATA_V4_TAG / ASSET).exists()

    # and with the release present, data-v4 serves ITS manifest-verified bytes
    v4_payload = b"v4-verified-bytes"
    manifest = make_data_v4_manifest({ASSET: v4_payload})

    def handler(url):
        if url == _url(DATA_V4_TAG, MANIFEST):
            return FakeResponse(200, manifest)
        if url == _url(DATA_V4_TAG, ASSET):
            return FakeResponse(200, v4_payload)
        raise AssertionError(f"unexpected url {url}")

    install_transport(monkeypatch, handler)
    fresh = build_workbook.release_client(DATA_V4_TAG, cache_dir=tmp_path, token_provider=no_token)
    data, prov = fresh.fetch(ASSET)

    assert data == v4_payload and data != v2_payload
    assert prov.tag == DATA_V4_TAG and prov.verified is True
    assert (tmp_path / "data-v2" / ASSET).read_bytes() == v2_payload  # untouched


# --- explicit data-v4 goes through manifest-verified loading -----------------


def test_explicit_data_v4_selection_exposes_manifest_verified_provenance(monkeypatch, tmp_path):
    payload = b"data-v4-payload"
    manifest = make_data_v4_manifest({ASSET: payload})

    def handler(url):
        if url == _url(DATA_V4_TAG, MANIFEST):
            return FakeResponse(200, manifest)
        if url == _url(DATA_V4_TAG, ASSET):
            return FakeResponse(200, payload)
        raise AssertionError(f"unexpected url {url}")

    install_transport(monkeypatch, handler)
    client = build_workbook.release_client(DATA_V4_TAG, cache_dir=tmp_path, token_provider=no_token)
    data, prov = client.fetch(ASSET)
    digest = hashlib.sha256(payload).hexdigest()

    assert data == payload
    assert prov.tag == DATA_V4_TAG  # active tag
    assert prov.verification == "publication_manifest_sha256"  # publication identity
    assert prov.verified is True and prov.expected_sha256 == digest  # manifest status
    assert prov.url == _url(DATA_V4_TAG, ASSET)  # source URL

    # the workbook surface exposes the same recorded Provenance per asset
    assert addin.fw_version(client) == DATA_V4_TAG
    frame = addin.fw_provenance(client)
    assert len(frame) == 1
    row = frame.iloc[0]
    assert row["tag"] == DATA_V4_TAG
    assert row["asset"] == ASSET
    assert row["url"] == _url(DATA_V4_TAG, ASSET)
    assert row["expected_sha256"] == digest
    assert bool(row["verified"]) is True
    assert row["verification"] == "publication_manifest_sha256"


def test_fw_load_constructs_a_fresh_client_when_the_tag_cell_changes():
    v2 = asyncio.run(addin.fw_load("data-v2"))
    v4 = asyncio.run(addin.fw_load(DATA_V4_TAG))

    assert v2.tag == "data-v2" and v4.tag == DATA_V4_TAG
    assert v4 is not v2
    assert v4.provenance_table() == []  # nothing remembered across tags
